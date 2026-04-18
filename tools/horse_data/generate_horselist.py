from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - only needed for JSON-only runs.
    load_workbook = None


TAG_HEADERS = [
    "SerialNumber",
    "Gender",
    "HorseId",
    "FactorFlg",
    "FactorName",
    "RareCd",
    "HeaderDetail",
    "Category",
    "Category_ht",
    "Paternal_t",
    "Paternal_tht",
    "Paternal_ht",
    "Paternal_hht",
    "Paternal_ttht",
    "Paternal_thht",
    "Paternal_htht",
    "Paternal_hhht",
    "Ped_t",
    "Ped_tt",
    "Ped_ttt",
    "Ped_tttt",
    "Ped_ttht",
    "Ped_tht",
    "Ped_thtt",
    "Ped_thht",
    "Ped_ht",
    "Ped_htt",
    "Ped_httt",
    "Ped_htht",
    "Ped_hht",
    "Ped_hhtt",
    "Ped_hhht",
    "Paternal_jik",
    "Paternal_mig",
    "Ped_All",
]


FACTOR_NAMES = {
    1: "短距離",
    2: "速力",
    3: "底力",
    4: "長距離",
    5: "適応力",
    6: "丈夫",
    7: "早熟",
    8: "晩成",
    9: "堅実",
    10: "気性難",
    11: "疾走",
    12: "中距離",
}
FACTOR_BADGE_LABELS = {
    "01": "短",
    "02": "速",
    "03": "底",
    "04": "長",
    "05": "適",
    "06": "丈",
    "07": "早",
    "08": "晩",
    "09": "堅",
    "10": "難",
    "11": "走",
    "12": "中",
}

SPEC_DETAILS = {1: "A", 2: "B", 3: "C"}
DIRT_DETAILS = {1: "△", 2: "〇", 3: "◎"}
BMS_RARE = {"X": "優", "W": "良", "V": "可", "U": "無印", "Y": "名牝"}
BMS_RARE_CD = {5: "X", 6: "W", 7: "V", 8: "U", 9: "Y"}
GENERATION_BY_SLOT = [2, 3, 4, 5, 5, 4, 5, 5, 3, 4, 5, 5, 4, 5, 5]
FACTOR_ORDER = [FACTOR_NAMES[index] for index in range(1, 13)]

PARENTAL_LINE_CODES = {
    "エクリプス系": "Ec",
    "フェアウェイ系": "Fa",
    "フェアトライアル系": "Fa",
    "オーエンテューダー系": "Ha",
    "オリオール系": "Ha",
    "カーレッド系": "Ha",
    "サンインロー系": "Ha",
    "ハイペリオン系": "Ha",
    "ハンプトン系": "Ha",
    "ファイントップ系": "Ha",
    "ロックフェラ系": "Ha",
    "クラリオン系": "He",
    "トウルビヨン系": "He",
    "ヘロド系": "He",
    "マイバブー系": "He",
    "ヒムヤー系": "Hi",
    "インテント系": "Ma",
    "マッチェム系": "Ma",
    "マンノウォー系": "Ma",
    "レリック系": "Ma",
    "エタン系": "Na",
    "ネイティヴダンサー系": "Na",
    "レイズアネイティヴ系": "Na",
    "ニアークティック系": "Ne",
    "ノーザンダンサー系": "Ne",
    "グレイソヴリン系": "Ns",
    "ゼダーン系": "Ns",
    "ソヴリンパス系": "Ns",
    "ナスルーラ系": "Ns",
    "ネヴァーセイダイ系": "Ns",
    "ネヴァーベンド系": "Ns",
    "フォルティノ系": "Ns",
    "プリンスリーギフト系": "Ns",
    "ボールドルーラー系": "Ns",
    "レッドゴッド系": "Ns",
    "ダンテ系": "Ph",
    "ネアルコ系": "Ph",
    "ファロス系": "Ph",
    "ファラリス系": "Ph",
    "ファリス系": "Ph",
    "モスボロー系": "Ph",
    "サーゲイロード系": "Ro",
    "ハビタット系": "Ro",
    "ヘイルトゥリーズン系": "Ro",
    "ロイヤルチャージャー系": "Ro",
    "セントサイモン系": "St",
    "プリンスキロ系": "St",
    "プリンスビオ系": "St",
    "プリンスローズ系": "St",
    "ボワルセル系": "St",
    "リボー系": "St",
    "ワイルドリスク系": "St",
    "スインフォード系": "Sw",
    "ブラントーム系": "Sw",
    "ブランドフォード系": "Sw",
    "ブレニム系": "Sw",
    "テディ系": "Te",
    "トムフール系": "To",
}


COL_GENDER = 1
COL_SERIAL_NUMBER = 2
COL_HORSE_ID = 3
COL_RARE = 4
COL_HORSE_NAME = 5
COL_PARENT_LINE = 6
COL_FACTOR_NAME_1 = 7
COL_FACTOR_NAME_2 = 8
COL_FACTOR_NAME_3 = 9
COL_ICON = 10
COL_DISTANCE_MIN = 11
COL_GROWTH = 13
COL_DIRT = 14
COL_HEALTH = 15
COL_CLEMENCY = 16
COL_RUNNING_STYLE = 17
COL_ACHIEVEMENT = 18
COL_POTENTIAL = 19
COL_STABLE = 20
COL_ABILITY = 21
COL_NAME_T = 23
COL_NAME_TT = 24
COL_NAME_TTT = 25
COL_NAME_TTTT = 26
COL_NAME_TTHT = 27
COL_NAME_THT = 28
COL_NAME_THTT = 29
COL_NAME_THHT = 30
COL_NAME_HT = 31
COL_NAME_HTT = 32
COL_NAME_HTTT = 33
COL_NAME_HTHT = 34
COL_NAME_HHT = 35
COL_NAME_HHTT = 36
COL_NAME_HHHT = 37
COL_PARENT_LINE_T = 38
COL_PARENT_LINE_TTHT = 42
COL_PARENT_LINE_THT = 43
COL_PARENT_LINE_THHT = 45
COL_PARENT_LINE_HT = 46
COL_PARENT_LINE_HTHT = 49
COL_PARENT_LINE_HHT = 50
COL_PARENT_LINE_HHHT = 52
COL_SON_T = 53
COL_SON_TT = 54
COL_SON_TTT = 55
COL_SON_TTTT = 56
COL_SON_TTHT = 57
COL_SON_THT = 58
COL_SON_THTT = 59
COL_SON_THHT = 60
COL_SON_HT = 61
COL_SON_HTT = 62
COL_SON_HTTT = 63
COL_SON_HTHT = 64
COL_SON_HHT = 65
COL_SON_HHTT = 66
COL_SON_HHHT = 67
COL_FACTOR_T1 = 68
COL_FACTOR_T2 = 69
COL_FACTOR_T3 = 70
COL_FACTOR_TT1 = 71
COL_FACTOR_TT2 = 72
COL_FACTOR_TT3 = 73
COL_FACTOR_TTT1 = 74
COL_FACTOR_TTT2 = 75
COL_FACTOR_TTT3 = 76
COL_FACTOR_TTTT1 = 77
COL_FACTOR_TTTT2 = 78
COL_FACTOR_TTTT3 = 79
COL_FACTOR_TTHT1 = 80
COL_FACTOR_TTHT2 = 81
COL_FACTOR_TTHT3 = 82
COL_FACTOR_THT1 = 83
COL_FACTOR_THT2 = 84
COL_FACTOR_THT3 = 85
COL_FACTOR_THTT1 = 86
COL_FACTOR_THTT2 = 87
COL_FACTOR_THTT3 = 88
COL_FACTOR_THHT1 = 89
COL_FACTOR_THHT2 = 90
COL_FACTOR_THHT3 = 91
COL_FACTOR_HT1 = 92
COL_FACTOR_HT2 = 93
COL_FACTOR_HT3 = 94
COL_FACTOR_HTT1 = 95
COL_FACTOR_HTT2 = 96
COL_FACTOR_HTT3 = 97
COL_FACTOR_HTTT1 = 98
COL_FACTOR_HTTT2 = 99
COL_FACTOR_HTTT3 = 100
COL_FACTOR_HTHT1 = 101
COL_FACTOR_HTHT2 = 102
COL_FACTOR_HTHT3 = 103
COL_FACTOR_HHT1 = 104
COL_FACTOR_HHT2 = 105
COL_FACTOR_HHT3 = 106
COL_FACTOR_HHTT1 = 107
COL_FACTOR_HHTT2 = 108
COL_FACTOR_HHTT3 = 109
COL_FACTOR_HHHT1 = 110
COL_FACTOR_HHHT2 = 111
COL_FACTOR_HHHT3 = 112

SOURCE_JSON_VERSION = 1
DEFAULT_SOURCE_JSON_PATH = Path("data/source/workbook.json")
SOURCE_HORSE_LIST_KEY = "horse_list"
SOURCE_ALL_KEY = "all"
SOURCE_SPECIAL_RARE_KEY = "special_rare"


FULLWIDTH_DIGIT_TABLE = str.maketrans("0123456789", "０１２３４５６７８９")
APOSTROPHE_BETWEEN_LETTERS = re.compile(r"([a-zA-Z])'([a-zA-Z])")


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value)


def digits(value: object) -> str:
    return re.sub(r"\D", "", text(value))


def fullwidth_digits(value: str) -> str:
    return value.translate(FULLWIDTH_DIGIT_TABLE)


def get_parent_line_name(parent_line: object) -> str:
    normalized = text(parent_line).replace("Nas", "Ns").replace("Nat", "Na")
    return normalized[:2]


def get_factor_codes(
    factor_url_1: object, factor_url_2: object, factor_url_3: object
) -> tuple[str, str, str]:
    factor_1 = ""
    factor_2 = ""
    factor_3 = ""

    value_1 = text(factor_url_1)
    value_2 = text(factor_url_2)
    value_3 = text(factor_url_3)

    if value_3:
        factor_1 = digits(value_1)
        factor_2 = digits(value_2)
        factor_3 = digits(value_3)
    elif value_2:
        factor_2 = digits(value_1)
        factor_3 = digits(value_2)
    elif value_1:
        factor_3 = digits(value_1)

    return factor_1, factor_2, factor_3


def get_factor_names(
    factor_url_1: object, factor_url_2: object, factor_url_3: object
) -> tuple[str, str, str]:
    return tuple(
        FACTOR_NAMES.get(int(code), "") if code else ""
        for code in get_factor_codes(factor_url_1, factor_url_2, factor_url_3)
    )


def get_factor_img(kind: str, factor_1: str, factor_2: str, factor_3: str) -> str:
    parts: list[str] = []
    count = 0

    if factor_2:
        parts.append(
            f'<td class="factor_{kind}" width="24"><img src="static/img/icn/icn_factor_{factor_2}.png" alt=""></td>'
        )
        count += 1

    if factor_3:
        parts.append(
            f'<td class="factor_{kind}" width="24"><img src="static/img/icn/icn_factor_{factor_3}.png" alt=""></td>'
        )
        count += 1

    if count == 1:
        parts.insert(0, f'<td class="factor_{kind}" width="24"></td>')
    elif count == 0:
        parts = [
            f'<td class="factor_{kind}" width="24"></td>',
            f'<td class="factor_{kind}" width="24"></td>',
        ]

    return "".join(parts)


def split_distance(value: object) -> tuple[str, str]:
    distance = text(value).replace("m", "")
    if not distance:
        return "", ""
    if "〜" in distance:
        return tuple(distance.split("〜", 1))
    if "～" in distance:
        return tuple(distance.split("～", 1))
    return distance, ""


def row_value(row: list[object], column: int) -> str:
    return text(row[column]) if column < len(row) else ""


def json_cell(value: object) -> object:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    return str(value)


def trim_trailing_blank_rows(rows: list[list[object]]) -> list[list[object]]:
    while rows and all(text(value) == "" for value in rows[-1]):
        rows.pop()
    return rows


def normalize_sheet_rows(rows: object, width: int) -> list[list[object]]:
    if not isinstance(rows, list):
        raise ValueError("source JSON sheet data must be a list of rows.")

    normalized: list[list[object]] = []
    for row in rows:
        if not isinstance(row, list):
            raise ValueError("source JSON sheet rows must be lists.")
        normalized.append((row[:width] + [None] * width)[:width])
    return normalized


def worksheet_rows(
    worksheet,
    *,
    min_row: int,
    max_col: int,
    stop_on_blank_column: int | None = None,
) -> list[list[object]]:
    rows: list[list[object]] = []
    for values in worksheet.iter_rows(
        min_row=min_row, max_col=max_col, values_only=True
    ):
        row = [json_cell(value) for value in values]
        if stop_on_blank_column is not None:
            stop_value = row[stop_on_blank_column - 1]
            if not text(stop_value):
                break
        rows.append(row)
    return trim_trailing_blank_rows(rows)


def load_excel_source(workbook_path: Path) -> dict[str, list[list[object]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsm sources.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return {
            SOURCE_HORSE_LIST_KEY: worksheet_rows(
                workbook["種牡馬一覧"], min_row=1, max_col=1
            ),
            SOURCE_ALL_KEY: worksheet_rows(
                workbook["ALL"],
                min_row=3,
                max_col=112,
                stop_on_blank_column=COL_SERIAL_NUMBER,
            ),
            SOURCE_SPECIAL_RARE_KEY: worksheet_rows(
                workbook["特別レア"], min_row=2, max_col=3
            ),
        }
    finally:
        workbook.close()


def load_source_json(source_json_path: Path) -> dict[str, list[list[object]]]:
    payload = json.loads(source_json_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("source JSON root must be an object.")

    sheets = payload.get("sheets", payload)
    if not isinstance(sheets, dict):
        raise ValueError("source JSON sheets must be an object.")

    missing_keys = [
        key
        for key in (SOURCE_HORSE_LIST_KEY, SOURCE_ALL_KEY, SOURCE_SPECIAL_RARE_KEY)
        if key not in sheets
    ]
    if missing_keys:
        raise ValueError(f"source JSON is missing sheets: {', '.join(missing_keys)}")

    return {
        SOURCE_HORSE_LIST_KEY: normalize_sheet_rows(
            sheets[SOURCE_HORSE_LIST_KEY], 1
        ),
        SOURCE_ALL_KEY: normalize_sheet_rows(sheets[SOURCE_ALL_KEY], 112),
        SOURCE_SPECIAL_RARE_KEY: normalize_sheet_rows(
            sheets[SOURCE_SPECIAL_RARE_KEY], 3
        ),
    }


def serialize_source_json(source: dict[str, list[list[object]]]) -> str:
    return json.dumps(
        {"version": SOURCE_JSON_VERSION, "sheets": source},
        ensure_ascii=False,
        separators=(",", ":"),
    )


def generate_source_json_text(workbook_path: Path) -> str:
    return serialize_source_json(load_excel_source(workbook_path))


def build_url_by_serial(rows: Iterable[list[object]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for index, cells in enumerate(rows, start=1):
        value = text(cells[0])
        if not value:
            continue
        match = re.search(r"/([0-9]+)\.html$", value)
        if match:
            mapping[f"{index:05d}"] = match.group(1)
    return mapping


def build_special_rare_sets(rows: Iterable[list[object]]) -> tuple[set[str], set[str]]:
    rare_8: set[str] = set()
    rare_7: set[str] = set()

    for cells in rows:
        if cells[0]:
            rare_8.add(digits(cells[0]))
        if len(cells) > 2 and cells[2]:
            rare_7.add(digits(cells[2]))

    rare_8.discard("")
    rare_7.discard("")
    return rare_8, rare_7


def iter_all_rows(rows: Iterable[list[object]]) -> Iterable[list[object]]:
    for values in rows:
        serial = text(values[COL_SERIAL_NUMBER - 1])
        if not serial:
            break
        yield [None, *values]


def build_factor_counts(row: list[object]) -> tuple[list[int], list[int], list[int]]:
    counter: Counter[str] = Counter()

    for factor_name in get_factor_names(
        row[COL_FACTOR_NAME_1], row[COL_FACTOR_NAME_2], row[COL_FACTOR_NAME_3]
    ):
        if factor_name and factor_name != FACTOR_NAMES[5]:
            counter[f"1{factor_name}"] += 1

    for slot_index, generation in enumerate(GENERATION_BY_SLOT):
        base = COL_FACTOR_T1 + (slot_index * 3)
        for factor_name in get_factor_names(row[base], row[base + 1], row[base + 2]):
            if factor_name and factor_name != FACTOR_NAMES[5]:
                counter[f"{generation}{factor_name}"] += 1

    counts_0 = [
        sum(counter[f"{generation}{factor_name}"] for generation in "12345")
        for factor_name in FACTOR_ORDER
    ]
    counts_1 = [
        sum(counter[f"{generation}{factor_name}"] for generation in "1234")
        for factor_name in FACTOR_ORDER
    ]
    counts_2 = [
        sum(counter[f"{generation}{factor_name}"] for generation in "123")
        for factor_name in FACTOR_ORDER
    ]
    return counts_0, counts_1, counts_2


def compute_horse_id(
    serial_number: str, row: list[object], url_by_serial: dict[str, str]
) -> str:
    recovered = digits(row[COL_HORSE_ID])
    if recovered:
        return recovered
    return url_by_serial.get(serial_number, "")


def compute_rare_cd(
    gender: str,
    horse_id: str,
    row: list[object],
    special_rare_8: set[str],
    special_rare_7: set[str],
) -> str:
    if gender == "0":
        raw_rare = text(row[COL_RARE])
        if raw_rare == "5":
            icon_code = digits(row[COL_ICON])
            if icon_code == "12":
                return "8"
            if icon_code == "11":
                return "7"
            if icon_code == "14":
                return "6"
            if horse_id in special_rare_8:
                return "8"
            if horse_id in special_rare_7:
                return "7"
            return "5"
        return raw_rare

    if text(row[COL_RARE]):
        icon_code = int(digits(row[COL_ICON]) or "0")
        return BMS_RARE_CD.get(icon_code, "")
    return "Z"


def build_header_detail(
    row: list[object],
    gender: str,
    horse_name: str,
    parent_line: str,
    factor_flag: str,
    factor_codes: tuple[str, str, str],
    rare_cd: str,
    counts_0: list[int],
    counts_1: list[int],
    counts_2: list[int],
) -> str:
    factor_1, factor_2, factor_3 = factor_codes
    ability = row_value(row, COL_ABILITY)
    running_style = row_value(row, COL_RUNNING_STYLE)[:1]
    growth = row_value(row, COL_GROWTH)[:1]
    achievement = SPEC_DETAILS.get(int(digits(row[COL_ACHIEVEMENT]) or "0"), "")
    clemency = SPEC_DETAILS.get(int(digits(row[COL_CLEMENCY]) or "0"), "")
    stable = SPEC_DETAILS.get(int(digits(row[COL_STABLE]) or "0"), "")
    potential = SPEC_DETAILS.get(int(digits(row[COL_POTENTIAL]) or "0"), "")
    health = SPEC_DETAILS.get(int(digits(row[COL_HEALTH]) or "0"), "")
    dirt = DIRT_DETAILS.get(int(digits(row[COL_DIRT]) or "0"), "")
    distance_min, distance_max = split_distance(row[COL_DISTANCE_MIN])

    parts: list[str] = []
    parts.append('<div class="horsedata2"><table class="horse_spec" width="100%"><tbody><tr>')

    if rare_cd == "8":
        parts.append('<th class="header01_8" width="10%">究極')
    elif rare_cd == "7":
        parts.append('<th class="header01_7" width="10%">究極')
    elif rare_cd == "6":
        parts.append('<th class="header01_6" width="10%">')
    elif rare_cd == "5":
        if digits(row[COL_ICON]) == "1057":
            parts.append('<th class="header01_F" width="10%">')
        else:
            parts.append(f'<th class="header01" width="10%">{fullwidth_digits(rare_cd)}')
    elif rare_cd == "Z":
        parts.append('<th class="header01_Z" width="10%">券')
    else:
        if gender == "0":
            parts.append(f'<th class="header01" width="10%">{fullwidth_digits(rare_cd)}')
        else:
            parts.append(f'<th class="header01" width="10%">{BMS_RARE.get(rare_cd, "")}')

    parts.append(f'</th><td width="60%"><label>{horse_name}<div class="factor_02_img" >')
    if factor_flag == "1":
        if factor_1:
            parts.append(f'<img src="static/img/icn/icn_factor_{factor_1}.png" alt="">')
        if factor_2:
            parts.append(f'<img src="static/img/icn/icn_factor_{factor_2}.png" alt="">')
        if factor_3:
            parts.append(f'<img src="static/img/icn/icn_factor_{factor_3}.png" alt="">')
    parts.append(f"&nbsp{parent_line}</div></label></td>")
    parts.append('<th class="header01" width="10%">非凡</th><td>')
    parts.append(ability if ability else "なし")
    parts.append("</td></tr></tbody></table>")

    parts.append('<table width="100%"><tbody><tr><th class="header01">')
    parts.append(
        "脚</th><th class=\"header01\">成</th> <th class=\"header01\">実</th><th class=\"header01\">気</th><th class=\"header01\">安</th>"
    )
    parts.append(
        '<th class="header01">底</th><th class="header01">体</th><th class="header01">適</th><th class="header01">距離</th>'
    )
    parts.append(
        '<th class="header01_f01">短</th><th class="header01_f02">速</th><th class="header01_f03">底</th><th class="header01_f04">長</th><th class="header01_f05">適</th>'
    )
    parts.append(
        '<th class="header01_f06">丈</th><th class="header01_f07">早</th><th class="header01_f08">晩</th><th class="header01_f09">堅</th><th class="header01_f10">難</th>'
    )
    parts.append('<th class="header01_f11">走</th><th class="header01_f12">中</th>')
    parts.append(f"</tr><tr><td>{running_style}</td><td>{growth}</td><td>{achievement}</td><td>{clemency}</td><td>{stable}</td>")
    parts.append(f"<td>{potential}</td><td>{health}</td><td>{dirt}</td><td>")
    if distance_min:
        parts.append(distance_min)
        parts.append("～")
        parts.append(distance_max)
    parts.append("</td>")
    for count in counts_0:
        parts.append(f"<td>{count:02d}</td>")
    parts.append("</tr></tbody></table>")

    parts.append('<table width="100%"><tbody><tr><th class="header01_01" colspan="12">１薄め</th><th class="header01_02" colspan="12">２薄め</th></tr><tr>')
    parts.append(
        '<th class="header01_f01">短</th><th class="header01_f02">速</th><th class="header01_f03">底</th><th class="header01_f04">長</th><th class="header01_f05">ダ</th>'
    )
    parts.append(
        '<th class="header01_f06">丈</th><th class="header01_f07">早</th><th class="header01_f08">晩</th><th class="header01_f09">堅</th><th class="header01_f10">難</th>'
    )
    parts.append(
        '<th class="header01_f11">走</th><th class="header01_f12">中</th><th class="header01_f01">短</th><th class="header01_f02">速</th><th class="header01_f03">底</th><th class="header01_f04">長</th><th class="header01_f05">ダ</th>'
    )
    parts.append(
        '<th class="header01_f06">丈</th><th class="header01_f07">早</th><th class="header01_f08">晩</th><th class="header01_f09">堅</th><th class="header01_f10">難</th>'
    )
    parts.append('<th class="header01_f11">走</th><th class="header01_f12">中</th></tr><tr>')
    for count in counts_1:
        parts.append(f"<td>{count:02d}</td>")
    for count in counts_2:
        parts.append(f"<td>{count:02d}</td>")
    parts.append("</tr></tbody></table></div>")
    return "".join(parts)


def compact_factor_codes(factor_codes: tuple[str, str, str]) -> list[str]:
    return [code for code in factor_codes if code]


def build_rare_badge(gender: str, rare_cd: str, icon_value: object) -> tuple[str, str]:
    icon_code = digits(icon_value)

    if rare_cd == "8":
        return "header01_8", "究極"
    if rare_cd == "7":
        return "header01_7", "究極"
    if rare_cd == "6":
        return "header01_6", ""
    if rare_cd == "5" and icon_code == "1057":
        return "header01_F", ""
    if rare_cd == "Z":
        return "header01_Z", "券"
    if gender == "0":
        return "header01", fullwidth_digits(rare_cd)
    return "header01", BMS_RARE.get(rare_cd, "")


def build_pedigree_entry(
    name: str,
    child_line: str,
    line_code: str,
    factor_codes: tuple[str, str, str],
) -> list[object]:
    return [name, child_line, line_code, compact_factor_codes(factor_codes)]


def build_records_from_source(
    source: dict[str, list[list[object]]]
) -> list[dict[str, object]]:
    url_by_serial = build_url_by_serial(source[SOURCE_HORSE_LIST_KEY])
    special_rare_8, special_rare_7 = build_special_rare_sets(
        source[SOURCE_SPECIAL_RARE_KEY]
    )

    records: list[dict[str, object]] = []

    for row in iter_all_rows(source[SOURCE_ALL_KEY]):
        serial_number = row_value(row, COL_SERIAL_NUMBER)
        gender = row_value(row, COL_GENDER)
        horse_id = compute_horse_id(serial_number, row, url_by_serial)

        self_factor_codes = get_factor_codes(
            row[COL_FACTOR_NAME_1], row[COL_FACTOR_NAME_2], row[COL_FACTOR_NAME_3]
        )
        self_factor_names = tuple(
            FACTOR_NAMES.get(int(code), "") if code else ""
            for code in self_factor_codes
        )
        factor_name = "".join(self_factor_names)
        factor_flag = "1" if factor_name else "0"
        rare_cd = compute_rare_cd(
            gender, horse_id, row, special_rare_8, special_rare_7
        )
        counts_0, counts_1, counts_2 = build_factor_counts(row)

        horse_name = row_value(row, COL_HORSE_NAME)
        parent_line = row_value(row, COL_PARENT_LINE)
        ability = row_value(row, COL_ABILITY)
        running_style = row_value(row, COL_RUNNING_STYLE)[:1]
        growth = row_value(row, COL_GROWTH)[:1]
        achievement = SPEC_DETAILS.get(int(digits(row[COL_ACHIEVEMENT]) or "0"), "")
        clemency = SPEC_DETAILS.get(int(digits(row[COL_CLEMENCY]) or "0"), "")
        stable = SPEC_DETAILS.get(int(digits(row[COL_STABLE]) or "0"), "")
        potential = SPEC_DETAILS.get(int(digits(row[COL_POTENTIAL]) or "0"), "")
        health = SPEC_DETAILS.get(int(digits(row[COL_HEALTH]) or "0"), "")
        dirt = DIRT_DETAILS.get(int(digits(row[COL_DIRT]) or "0"), "")
        distance_min, distance_max = split_distance(row[COL_DISTANCE_MIN])
        rare_badge_class, rare_badge_label = build_rare_badge(
            gender, rare_cd, row[COL_ICON]
        )
        category_ht = row_value(row, COL_SON_HT)

        paternal_t = PARENTAL_LINE_CODES.get(parent_line.strip(), "")
        paternal_t_line = get_parent_line_name(row[COL_PARENT_LINE_T])
        paternal_tht = get_parent_line_name(row[COL_PARENT_LINE_THT])
        paternal_ht = get_parent_line_name(row[COL_PARENT_LINE_HT])
        paternal_hht = get_parent_line_name(row[COL_PARENT_LINE_HHT])
        paternal_ttht = get_parent_line_name(row[COL_PARENT_LINE_TTHT])
        paternal_thht = get_parent_line_name(row[COL_PARENT_LINE_THHT])
        paternal_htht = get_parent_line_name(row[COL_PARENT_LINE_HTHT])
        paternal_hhht = get_parent_line_name(row[COL_PARENT_LINE_HHHT])

        paternal_jik = ""
        paternal_mig = ""
        paternal_name = f"[自身{horse_name}]{parent_line}"
        pedigree: list[list[object]] = []

        pedigree: list[list[object]] = []

        factor_t = get_factor_codes(row[COL_FACTOR_T1], row[COL_FACTOR_T2], row[COL_FACTOR_T3])
        if gender == "0":
            ped_t = (
                '<tr><td align="center" class="father_0" width="15">父</td>'
                f'<td colspan="4" class="omoshiro_0">{row_value(row, COL_NAME_T)}</td>'
                f'<td width="180" class="factor_02_img omoshiro">{row_value(row, COL_SON_T)}</td>'
                f'<td width="50" class="omoshiro_11">{get_parent_line_name(row[COL_PARENT_LINE_T])}</td>'
                f"{get_factor_img('omoshiro', *factor_t)}</tr>"
            )
        else:
            ped_t = (
                '<tr><td align="center" class="father_0" width="15">父</td>'
                f'<td colspan="4" class="omoshiro_mare_0">{row_value(row, COL_NAME_T)}</td>'
                f'<td width="180" class="factor_02_img omoshiro_mare_11">{row_value(row, COL_SON_T)}</td>'
                f'<td width="50" class="omoshiro_mare_12">{get_parent_line_name(row[COL_PARENT_LINE_T])}</td>'
                f"{get_factor_img('omoshiro_mare', *factor_t)}</tr>"
            )
        paternal_name += f"[１父{row_value(row, COL_NAME_T)}]"

        factor_tt = get_factor_codes(
            row[COL_FACTOR_TT1], row[COL_FACTOR_TT2], row[COL_FACTOR_TT3]
        )
        ped_tt = (
            '<tr><td align="center" class="father_1" rowspan="7" width="15"></td>'
            '<td class="father_0" width="15">父</td>'
            f'<td colspan="3" class="horse_0">{row_value(row, COL_NAME_TT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_TT)}</td>'
            '<td width="50" class="horse_1"></td>'
            f"{get_factor_img('horse', *factor_tt)}</tr>"
        )
        paternal_name += f"[父父{row_value(row, COL_NAME_TT)}]"

        factor_ttt = get_factor_codes(
            row[COL_FACTOR_TTT1], row[COL_FACTOR_TTT2], row[COL_FACTOR_TTT3]
        )
        ped_ttt = (
            '<tr><td align="center" class="father_1" rowspan="3" width="15"></td>'
            '<td class="father_0" width="15">父</td>'
            f'<td colspan="2" class="horse_0">{row_value(row, COL_NAME_TTT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_TTT)}</td>'
            '<td width="50" class="horse_1"></td>'
            f"{get_factor_img('horse', *factor_ttt)}</tr>"
        )
        paternal_name += f"[以外{row_value(row, COL_NAME_TTT)}]"

        factor_tttt = get_factor_codes(
            row[COL_FACTOR_TTTT1], row[COL_FACTOR_TTTT2], row[COL_FACTOR_TTTT3]
        )
        ped_tttt = (
            '<tr><td align="center" class="father_1" width="15"></td>'
            '<td class="father" width="15">父</td>'
            f'<td class="horse_0">{row_value(row, COL_NAME_TTTT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_TTTT)}</td>'
            '<td width="50" class="horse_1"></td>'
            f"{get_factor_img('horse', *factor_tttt)}</tr>"
        )
        paternal_name += f"[以外{row_value(row, COL_NAME_TTTT)}]"

        factor_ttht = get_factor_codes(
            row[COL_FACTOR_TTHT1], row[COL_FACTOR_TTHT2], row[COL_FACTOR_TTHT3]
        )
        if gender == "0":
            ped_ttht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="migoto">{row_value(row, COL_NAME_TTHT)}</td>'
                f'<td width="180" class="factor_02_img migoto_0">{row_value(row, COL_SON_TTHT)}</td>'
                f'<td width="50" class="migoto_1">{paternal_ttht}</td>'
                f"{get_factor_img('migoto', *factor_ttht)}</tr>"
            )
            paternal_name += f"[見事{row_value(row, COL_NAME_TTHT)}]"
        else:
            ped_ttht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="horse_0">{row_value(row, COL_NAME_TTHT)}</td>'
                f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_TTHT)}</td>'
                '<td width="50" class="horse_1"></td>'
                f"{get_factor_img('horse', *factor_ttht)}</tr>"
            )
            paternal_name += f"[以外{row_value(row, COL_NAME_TTHT)}]"
        paternal_mig += paternal_ttht

        factor_tht = get_factor_codes(
            row[COL_FACTOR_THT1], row[COL_FACTOR_THT2], row[COL_FACTOR_THT3]
        )
        if gender == "0":
            ped_tht = (
                '<tr><td class="mother_0">母</td><td class="father_0" rowspan="1">父</td>'
                f'<td colspan="2" class="omoshiro_0">{row_value(row, COL_NAME_THT)}</td>'
                f'<td width="180" class="factor_02_img omoshiro">{row_value(row, COL_SON_THT)}</td>'
                f'<td width="50" class="omoshiro_2">{paternal_tht}</td>'
                f"{get_factor_img('omoshiro', *factor_tht)}</tr>"
            )
        else:
            ped_tht = (
                '<tr><td class="mother_0">母</td><td class="father_0" rowspan="1">父</td>'
                f'<td colspan="2" class="omoshiro_mare_0">{row_value(row, COL_NAME_THT)}</td>'
                f'<td width="180" class="factor_02_img omoshiro_mare_11">{row_value(row, COL_SON_THT)}</td>'
                f'<td width="50" class="omoshiro_mare_2">{paternal_tht}</td>'
                f"{get_factor_img('omoshiro_mare', *factor_tht)}</tr>"
            )
        paternal_name += f"[１薄{row_value(row, COL_NAME_THT)}]"
        paternal_jik += paternal_tht

        factor_thtt = get_factor_codes(
            row[COL_FACTOR_THTT1], row[COL_FACTOR_THTT2], row[COL_FACTOR_THTT3]
        )
        ped_thtt = (
            '<tr><td class="mother_1" rowspan="2"></td><td class="father_1"></td>'
            '<td class="father">父</td>'
            f'<td class="horse_0">{row_value(row, COL_NAME_THTT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_THTT)}</td>'
            '<td width="50" class="horse_1"></td>'
            f"{get_factor_img('horse', *factor_thtt)}</tr>"
        )
        paternal_name += f"[以外{row_value(row, COL_NAME_THTT)}]"

        factor_thht = get_factor_codes(
            row[COL_FACTOR_THHT1], row[COL_FACTOR_THHT2], row[COL_FACTOR_THHT3]
        )
        if gender == "0":
            ped_thht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="migoto">{row_value(row, COL_NAME_THHT)}</td>'
                f'<td width="180" class="factor_02_img migoto_0">{row_value(row, COL_SON_THHT)}</td>'
                f'<td width="50" class="migoto_1">{get_parent_line_name(paternal_thht)}</td>'
                f"{get_factor_img('migoto', *factor_thht)}</tr>"
            )
            paternal_name += f"[見事{row_value(row, COL_NAME_THHT)}]"
            paternal_mig += paternal_thht
        else:
            ped_thht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="horse_0">{row_value(row, COL_NAME_THHT)}</td>'
                f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_THHT)}</td>'
                '<td width="50" class="horse_1"></td>'
                f"{get_factor_img('horse', *factor_thht)}</tr>"
            )
            paternal_name += f"[以外{row_value(row, COL_NAME_THHT)}]"

        factor_ht = get_factor_codes(
            row[COL_FACTOR_HT1], row[COL_FACTOR_HT2], row[COL_FACTOR_HT3]
        )
        if gender == "0":
            ped_ht = (
                '<tr><td class="mother_0">母</td><td class="father_0">父</td>'
                f'<td colspan="3" class="omoshiro_0">{row_value(row, COL_NAME_HT)}</td>'
                f'<td width="180" class="factor_02_img omoshiro">{row_value(row, COL_SON_HT)}</td>'
                f'<td width="50" class="omoshiro_12">{paternal_ht}</td>'
                f"{get_factor_img('omoshiro', *factor_ht)}</tr>"
            )
        else:
            ped_ht = (
                '<tr><td class="mother_0">母</td><td class="father_0">父</td>'
                f'<td colspan="3" class="omoshiro_mare_0">{row_value(row, COL_NAME_HT)}</td>'
                f'<td width="180" class="factor_02_img omoshiro_mare_11">{row_value(row, COL_SON_HT)}</td>'
                f'<td width="50" class="omoshiro_mare_12">{paternal_ht}</td>'
                f"{get_factor_img('omoshiro_mare', *factor_ht)}</tr>"
            )
        paternal_name += f"[母父{row_value(row, COL_NAME_HT)}]"

        factor_htt = get_factor_codes(
            row[COL_FACTOR_HTT1], row[COL_FACTOR_HTT2], row[COL_FACTOR_HTT3]
        )
        ped_htt = (
            '<tr><td class="mother_1" rowspan="6"></td><td class="father_1" rowspan="3"></td>'
            '<td class="father_0">父</td>'
            f'<td colspan="2" class="horse_0">{row_value(row, COL_NAME_HTT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_HTT)}</td>'
            '<td width="50" class="horse_1"></td>'
            f"{get_factor_img('horse', *factor_htt)}</tr>"
        )
        paternal_name += f"[以外{row_value(row, COL_NAME_HTT)}]"

        factor_httt = get_factor_codes(
            row[COL_FACTOR_HTTT1], row[COL_FACTOR_HTTT2], row[COL_FACTOR_HTTT3]
        )
        ped_httt = (
            '<tr><td class="father_1"></td><td class="father">父</td>'
            f'<td class="horse_0">{row_value(row, COL_NAME_HTTT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_HTTT)}</td>'
            '<td width="50" class="horse_1"></td>'
            f"{get_factor_img('horse', *factor_httt)}</tr>"
        )
        paternal_name += f"[以外{row_value(row, COL_NAME_HTTT)}]"

        factor_htht = get_factor_codes(
            row[COL_FACTOR_HTHT1], row[COL_FACTOR_HTHT2], row[COL_FACTOR_HTHT3]
        )
        if gender == "0":
            ped_htht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="migoto">{row_value(row, COL_NAME_HTHT)}</td>'
                f'<td width="180" class="factor_02_img migoto_0">{row_value(row, COL_SON_HTHT)}</td>'
                f'<td width="50" class="migoto_1">{paternal_htht}</td>'
                f"{get_factor_img('migoto', *factor_htht)}</tr>"
            )
            paternal_name += f"[見事{row_value(row, COL_NAME_HTHT)}]"
        else:
            ped_htht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="horse_0">{row_value(row, COL_NAME_HTHT)}</td>'
                f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_HTHT)}</td>'
                '<td width="50" class="horse_1"></td>'
                f"{get_factor_img('horse', *factor_htht)}</tr>"
            )
            paternal_name += f"[以外{row_value(row, COL_NAME_HTHT)}]"
        paternal_mig += paternal_htht

        factor_hht = get_factor_codes(
            row[COL_FACTOR_HHT1], row[COL_FACTOR_HHT2], row[COL_FACTOR_HHT3]
        )
        if gender == "0":
            ped_hht = (
                '<tr><td class="mother_0">母</td><td class="father_0">父</td>'
                f'<td colspan="2" class="omoshiro_0">{row_value(row, COL_NAME_HHT)}</td>'
                f'<td width="180" class="factor_02_img omoshiro">{row_value(row, COL_SON_HHT)}</td>'
                f'<td  width="50" class="omoshiro_2">{paternal_hht}</td>'
                f"{get_factor_img('omoshiro', *factor_hht)}</tr>"
            )
        else:
            ped_hht = (
                '<tr><td class="mother_0">母</td><td class="father_0">父</td>'
                f'<td colspan="2" class="omoshiro_mare_0">{row_value(row, COL_NAME_HHT)}</td>'
                f'<td width="180" class="factor_02_img omoshiro_mare_11">{row_value(row, COL_SON_HHT)}</td>'
                f'<td  width="50" class="omoshiro_mare_2">{paternal_hht}</td>'
                f"{get_factor_img('omoshiro_mare', *factor_hht)}</tr>"
            )
        paternal_name += f"[１薄{row_value(row, COL_NAME_HHT)}]"
        paternal_jik += paternal_hht

        factor_hhtt = get_factor_codes(
            row[COL_FACTOR_HHTT1], row[COL_FACTOR_HHTT2], row[COL_FACTOR_HHTT3]
        )
        ped_hhtt = (
            '<tr><td class="mother_1" rowspan="2"></td><td class="father_1" rowspan="1"></td>'
            '<td class="father">父</td>'
            f'<td class="horse_0">{row_value(row, COL_NAME_HHTT)}</td>'
            f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_HHTT)}</td>'
            '<td class="horse_1" width="50"></td>'
            f"{get_factor_img('horse', *factor_hhtt)}</tr>"
        )
        paternal_name += f"[以外{row_value(row, COL_NAME_HHTT)}]"

        factor_hhht = get_factor_codes(
            row[COL_FACTOR_HHHT1], row[COL_FACTOR_HHHT2], row[COL_FACTOR_HHHT3]
        )
        if gender == "0":
            ped_hhht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="migoto">{row_value(row, COL_NAME_HHHT)}</td>'
                f'<td width="180" class="factor_02_img migoto_0">{row_value(row, COL_SON_HHHT)}</td>'
                f'<td width="50" class="migoto_1">{paternal_hhht}</td>'
                f"{get_factor_img('migoto', *factor_hhht)}</tr>"
            )
            paternal_name += f"[見事{row_value(row, COL_NAME_HHHT)}]"
        else:
            ped_hhht = (
                '<tr><td class="mother">母</td><td class="father">父</td>'
                f'<td class="horse_0">{row_value(row, COL_NAME_HHHT)}</td>'
                f'<td width="180" class="factor_02_img horse_0">{row_value(row, COL_SON_HHHT)}</td>'
                '<td width="50" class="horse_1"></td>'
                f"{get_factor_img('horse', *factor_hhht)}</tr>"
            )
            paternal_name += f"[以外{row_value(row, COL_NAME_HHHT)}]"
        paternal_mig += paternal_hhht

        pedigree = [
            build_pedigree_entry(
                row_value(row, COL_NAME_T),
                row_value(row, COL_SON_T),
                paternal_t_line,
                factor_t,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_TT),
                row_value(row, COL_SON_TT),
                "",
                factor_tt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_TTT),
                row_value(row, COL_SON_TTT),
                "",
                factor_ttt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_TTTT),
                row_value(row, COL_SON_TTTT),
                "",
                factor_tttt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_TTHT),
                row_value(row, COL_SON_TTHT),
                paternal_ttht if gender == "0" else "",
                factor_ttht,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_THT),
                row_value(row, COL_SON_THT),
                paternal_tht,
                factor_tht,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_THTT),
                row_value(row, COL_SON_THTT),
                "",
                factor_thtt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_THHT),
                row_value(row, COL_SON_THHT),
                paternal_thht if gender == "0" else "",
                factor_thht,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HT),
                category_ht,
                paternal_ht,
                factor_ht,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HTT),
                row_value(row, COL_SON_HTT),
                "",
                factor_htt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HTTT),
                row_value(row, COL_SON_HTTT),
                "",
                factor_httt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HTHT),
                row_value(row, COL_SON_HTHT),
                paternal_htht if gender == "0" else "",
                factor_htht,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HHT),
                row_value(row, COL_SON_HHT),
                paternal_hht,
                factor_hht,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HHTT),
                row_value(row, COL_SON_HHTT),
                "",
                factor_hhtt,
            ),
            build_pedigree_entry(
                row_value(row, COL_NAME_HHHT),
                row_value(row, COL_SON_HHHT),
                paternal_hhht if gender == "0" else "",
                factor_hhht,
            ),
        ]

        record = {
            "SerialNumber": serial_number,
            "Gender": gender,
            "HorseId": horse_id,
            "FactorFlg": factor_flag,
            "RareCd": rare_cd,
            "Category": parent_line,
            "Category_ht": category_ht,
            "Paternal_t": paternal_t,
            "Paternal_ht": paternal_ht,
            "Paternal_jik": paternal_jik,
            "Paternal_mig": paternal_mig,
            "Ped_All": paternal_name + "".join(self_factor_names),
            "card": {
                "name": horse_name,
                "ability": ability,
                "rareBadgeClass": rare_badge_class,
                "rareBadgeLabel": rare_badge_label,
                "selfFactorCodes": compact_factor_codes(self_factor_codes),
                "stats": {
                    "runningStyle": running_style,
                    "growth": growth,
                    "achievement": achievement,
                    "clemency": clemency,
                    "stable": stable,
                    "potential": potential,
                    "health": health,
                    "dirt": dirt,
                    "distanceMin": distance_min,
                    "distanceMax": distance_max,
                },
                "factorCounts": [counts_0, counts_1, counts_2],
                "pedigree": pedigree,
            },
        }
        records.append(record)

    return records


def build_records(workbook_path: Path) -> list[dict[str, object]]:
    return build_records_from_source(load_excel_source(workbook_path))


def build_records_from_json(source_json_path: Path) -> list[dict[str, object]]:
    return build_records_from_source(load_source_json(source_json_path))


def serialize_records(records: list[dict[str, object]]) -> str:
    return json.dumps(records, ensure_ascii=False, separators=(",", ":"))


def factor_sort_key(name: str) -> tuple[int, str]:
    return (0 if name[:1].isascii() else 1, name.casefold())


def build_factor_options(records: list[dict[str, object]]) -> list[dict[str, object]]:
    factor_map: dict[str, set[str]] = {}

    for record in records:
        card = record.get("card")
        if not isinstance(card, dict):
            continue

        pedigree = card.get("pedigree")
        if not isinstance(pedigree, list):
            continue

        for entry in pedigree:
            if not isinstance(entry, list) or len(entry) < 4:
                continue

            name = text(entry[0]).strip()
            factor_codes = entry[3]

            if not name or not isinstance(factor_codes, list):
                continue

            normalized_codes = {
                text(code).strip()
                for code in factor_codes
                if text(code).strip()
            }

            if normalized_codes:
                factor_map.setdefault(name, set()).update(normalized_codes)

    return [
        {
            "name": name,
            "id": name,
            "badges": [
                FACTOR_BADGE_LABELS[code]
                for code in sorted(factor_map[name])
                if code in FACTOR_BADGE_LABELS
            ],
        }
        for name in sorted(factor_map, key=factor_sort_key)
    ]


def serialize_factor_options(factor_options: list[dict[str, object]]) -> str:
    return json.dumps(
        {"factor": factor_options},
        ensure_ascii=False,
        separators=(",", ":"),
    )


def generate_horselist_text(workbook_path: Path) -> str:
    return serialize_records(build_records(workbook_path))


def generate_horselist_text_from_json(source_json_path: Path) -> str:
    return serialize_records(build_records_from_json(source_json_path))


def generate_factor_text(workbook_path: Path) -> str:
    return serialize_factor_options(build_factor_options(build_records(workbook_path)))


def generate_factor_text_from_json(source_json_path: Path) -> str:
    return serialize_factor_options(
        build_factor_options(build_records_from_json(source_json_path))
    )


def write_text(path: Path, text_value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(text_value)


def default_workbook_path(cwd: Path) -> Path:
    matches = sorted(cwd.glob("*.xlsm"))
    if not matches:
        raise FileNotFoundError("No .xlsm workbook found in the current directory.")
    return matches[0]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate horselist.json and factor.json from workbook or source JSON data."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Workbook to read. Defaults to the first .xlsm in the current directory when source JSON is not used.",
    )
    parser.add_argument(
        "--source-json",
        type=Path,
        default=None,
        help=f"Source JSON to read. Defaults to {DEFAULT_SOURCE_JSON_PATH} when no .xlsm workbook is found.",
    )
    parser.add_argument(
        "--export-source-json",
        type=Path,
        default=None,
        help="Write a source JSON snapshot from the workbook before generating outputs.",
    )
    parser.add_argument(
        "--output",
        action="append",
        type=Path,
        default=None,
        help="Output path. Repeat to write the same content to multiple files.",
    )
    parser.add_argument(
        "--factor-output",
        action="append",
        type=Path,
        default=None,
        help="factor.json output path. Repeat to write the same content to multiple files.",
    )
    return parser.parse_args()


def load_records_for_args(
    args: argparse.Namespace, cwd: Path
) -> tuple[list[dict[str, object]], Path]:
    if args.source_json is not None:
        return build_records_from_json(args.source_json), cwd

    if args.workbook is not None:
        return build_records(args.workbook), args.workbook.parent

    matches = sorted(cwd.glob("*.xlsm"))
    if matches:
        workbook_path = matches[0]
        return build_records(workbook_path), workbook_path.parent

    default_source_json = cwd / DEFAULT_SOURCE_JSON_PATH
    if default_source_json.exists():
        return build_records_from_json(default_source_json), cwd

    raise FileNotFoundError(
        f"No .xlsm workbook found and {DEFAULT_SOURCE_JSON_PATH} does not exist."
    )


def main() -> int:
    args = parse_args()
    cwd = Path.cwd()

    if args.export_source_json is not None:
        workbook_path = args.workbook or default_workbook_path(cwd)
        write_text(args.export_source_json, generate_source_json_text(workbook_path))

    records, default_output_root = load_records_for_args(args, cwd)
    output_paths = args.output or [default_output_root / "json" / "horselist.json"]
    factor_output_paths = args.factor_output or [
        default_output_root / "json" / "factor.json"
    ]
    text_value = serialize_records(records)
    factor_text_value = serialize_factor_options(build_factor_options(records))

    for output_path in output_paths:
        write_text(output_path, text_value)

    for output_path in factor_output_paths:
        write_text(output_path, factor_text_value)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
